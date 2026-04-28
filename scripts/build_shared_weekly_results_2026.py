from __future__ import annotations

import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from project_paths import SHARED_OVERVIEW_FILE, WEEKLY_RESULTS_FILE


INPUT_FILE = WEEKLY_RESULTS_FILE
OUTPUT_FILE = SHARED_OVERVIEW_FILE

TITLE_FILL = PatternFill("solid", fgColor="A61E22")
TITLE_FONT = Font(color="FFFFFF", bold=True, size=16)
SUBTITLE_FONT = Font(color="666666", italic=True, size=10)
CARD_FILL = PatternFill("solid", fgColor="F4E7E8")
HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
HEADER_FONT = Font(bold=True, color="1F1F1F")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

DISTANCE_ORDER = {
    "800 m": 1,
    "1500 m": 2,
    "3000 m": 3,
    "5 km": 4,
    "10 km": 5,
    "Halvmaraton": 6,
    "Maraton": 7,
    "42 km": 7,
    "30 km": 8,
    "60 km": 9,
}

EVENT_NAME_OVERRIDES = {
    "Bislett DS 5k": "Bislett distanseserie 2- 5K",
    "Bislett DS 10k": "Bislett distanseserie 2- 10K",
    "Sparebank 1 S?R-NORGE Drammen10K": "Drammen10K",
    "Sparebank 1 SØR-NORGE Drammen10K": "Drammen10K",
    "Fredrikstadl?pet": "Fredrikstadløpet",
    "Holmestrand maraton 5k": "Holmestrand maraton",
    "NM terrengl?p kort l?ype": "NM terrengløp kort løype",
    "Oslo L?psfestival - 5'ern v?r!": "Oslo Løpsfestival - 5'ern vår!",
}

TEXT_REPLACEMENTS = {
    "?dne Andersen Andersen": "Ådne Andersen",
    "Ådne Andersen Andersen": "Ådne Andersen",
    "Anna Marie Sirev?g": "Anna Marie Sirevåg",
    "Hege Njå Bjørkmanm": "Hege Njå Bjørkmann",
    "Kasper S-R": "Kasper Sørlie-Reininger",
    "Madel?ne Holum": "Madelène Holum",
    "Madel?ne Wanvik Holum": "Madelène Wanvik Holum",
}

NOTE_REPLACEMENTS = {
    ":first place medal:": "1. plass",
    "first place medal": "1. plass",
    ":second place medal:": "2. plass",
    "second place medal": "2. plass",
    ":third place medal:": "3. plass",
    "third place medal": "3. plass",
    "rabbit": "fartsholder",
    ":probing cane:": "para",
    ":star struck:": "",
    ":star:": "",
    "â­": "⭐",
}


def repair_mojibake(text: str) -> str:
    repaired = text
    for _ in range(2):
        if not any(marker in repaired for marker in ("Ã", "Â", "â")):
            break
        try:
            candidate = repaired.encode("latin1").decode("utf-8")
        except Exception:
            break
        if candidate == repaired:
            break
        repaired = candidate
    return repaired


def normalize_display_text(value: object) -> str:
    text = repair_mojibake(str(value or ""))
    for old, new in TEXT_REPLACEMENTS.items():
        text = text.replace(old, new)
    return text


def clean_note(value: object) -> str:
    text = normalize_display_text(value).strip()
    if not text or text.lower() == "nan":
        return ""

    lowered = text.casefold()
    if ("helgens" in lowered or "trener" in lowered) and ("rabbit" in lowered or "fartsholder" in lowered):
        return "fartsholder"

    for old, new in NOTE_REPLACEMENTS.items():
        text = text.replace(old, new)

    text = re.sub(r"<@[^>]+>", "", text)
    text = text.replace("︎", "")
    text = re.sub(r":[^:\s]+:", " ", text)
    text = re.sub(r"\bstar struck\b", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\bstar\b", "⭐", text, flags=re.IGNORECASE)
    text = re.sub(r"\bnew\b", "", text, flags=re.IGNORECASE)
    text = re.sub(r"^\.\s*", "", text)
    text = re.sub(r"^\w*evåg\s*:\s*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\(([^)]*)\)", r"\1", text)
    text = re.sub(r"\s+", " ", text)

    text = text.replace("Vinner av F65 klassen 1. plass", "Vinner av F65-klassen, 1. plass")
    text = text.replace("Vinner av F65 klassen1. plass", "Vinner av F65-klassen, 1. plass")
    text = text.replace("Vinner av F65 klassen", "Vinner av F65-klassen")
    text = text.replace("Innendørs PB", "innendørs PB")
    text = text.replace("1. plass M18 19", "1. plass, M18-19")
    text = text.replace("1. plass, M18 19", "1. plass, M18-19")
    text = text.replace("3. plass innendørs PB", "3. plass, innendørs PB")
    text = text.replace("4. plass ⭐", "4. plass")
    text = text.replace("4. plass:⭐", "4. plass")
    text = text.replace("para para", "para")
    text = text.replace("M75 79", "M75-79")
    text = text.replace("Offensiv åpning på .", "Offensiv åpning.")

    text = text.replace(":", ", ")
    text = re.sub(r",\s*,+", ", ", text)
    text = re.sub(r"\s*,\s*", ", ", text)
    text = re.sub(r"\s+", " ", text)
    text = text.strip(" .,:;()")

    if re.fullmatch(r"[KM]\s*pos,\s*\d+", text, flags=re.IGNORECASE):
        text = text.replace(",", ":")

    if text in {"@", "()", "( )", "en", "new", "⭐"}:
        return ""
    return text


def extract_place(position: object, note: object) -> str:
    if position is not None and not pd.isna(position):
        try:
            return str(int(position))
        except Exception:
            return str(position).strip()

    note_text = str(note or "").lower()
    if "first place" in note_text:
        return "1"
    if "second place" in note_text:
        return "2"
    if "third place" in note_text:
        return "3"

    match = re.search(r"(\d+)\.\s*plass", note_text)
    if match:
        return match.group(1)

    match = re.search(r"pos:\s*(\d+)", note_text)
    if match:
        return match.group(1)

    return ""


def format_date(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    try:
        return pd.to_datetime(value).strftime("%d.%m.%Y")
    except Exception:
        return str(value).strip()


def parse_time_for_sort(value: object) -> float:
    if value is None or pd.isna(value):
        return float("inf")

    text = str(value).strip().replace(",", ".")
    if not text or text.lower() == "nan":
        return float("inf")

    try:
        if ":" not in text and text.count(".") >= 2:
            parts = text.split(".")
            if len(parts) == 3:
                minutes = int(parts[0])
                seconds = int(parts[1])
                hundredths = int(parts[2])
                return minutes * 60 + seconds + hundredths / 100

        total = 0.0
        for part in text.split(":"):
            total = total * 60 + float(part)
        return total
    except Exception:
        return float("inf")


def build_display_rows(df: pd.DataFrame) -> pd.DataFrame:
    working = df.copy()
    working = working[working["athlete_name"].notna()].copy()

    gender_column = "gender" if "gender" in working.columns else "WA Kjønn"
    class_column = "class_name" if "class_name" in working.columns else "category"

    working["published_date_sort"] = pd.to_datetime(working["published_date"], errors="coerce")
    working["week_sort"] = pd.to_numeric(working["week_number"], errors="coerce")
    working["Uke"] = working["week_number"].apply(lambda value: f"Uke {int(value)}" if pd.notna(value) else "")
    working["Dato"] = working["published_date"].apply(format_date)
    working["Løp"] = (
        working["event_name"]
        .fillna("")
        .astype(str)
        .map(normalize_display_text)
        .str.strip()
        .replace(EVENT_NAME_OVERRIDES)
    )
    working["Navn"] = working["athlete_name"].fillna("").astype(str).map(normalize_display_text).str.strip()
    working["Kjønn"] = working[gender_column].fillna("").astype(str).map(normalize_display_text).str.strip()
    working["Klasse"] = working[class_column].fillna("").astype(str).map(normalize_display_text).str.strip()
    working["Distanse"] = working["distance"].fillna("").astype(str).map(normalize_display_text).str.strip()
    working["Tid"] = (
        working["result_time_normalized"]
        .fillna(working["result_time_raw"])
        .fillna("")
        .astype(str)
        .str.strip()
        .replace("nan", "")
    )
    working["Plass"] = [extract_place(position, note) for position, note in zip(working["position"], working["notes"])]
    working["Kort note"] = working["notes"].apply(clean_note)
    working["distance_sort"] = working["Distanse"].map(DISTANCE_ORDER).fillna(99)
    working["time_sort"] = working["Tid"].apply(parse_time_for_sort)

    display = working[
        [
            "week_sort",
            "published_date_sort",
            "distance_sort",
            "time_sort",
            "Uke",
            "Dato",
            "Løp",
            "Navn",
            "Kjønn",
            "Klasse",
            "Distanse",
            "Tid",
            "Plass",
            "Kort note",
        ]
    ].copy()
    display = display.sort_values(
        ["week_sort", "published_date_sort", "Løp", "distance_sort", "time_sort", "Navn"],
        ascending=[False, False, True, True, True, True],
        na_position="last",
    )
    display = display.drop(columns=["week_sort", "published_date_sort", "distance_sort", "time_sort"])
    return display


def build_week_summary(df: pd.DataFrame) -> pd.DataFrame:
    working = df.copy()
    working = working[working["athlete_name"].notna()].copy()
    grouped = (
        working.groupby(["week_number"], dropna=False)
        .agg(
            published_date=("published_date", "max"),
            resultater=("athlete_name", "count"),
            lopere=("athlete_name", "nunique"),
            lop=(
                "event_name",
                lambda values: ", ".join(
                    sorted(
                        {
                            EVENT_NAME_OVERRIDES.get(str(value).strip(), str(value).strip())
                            for value in values
                            if pd.notna(value)
                        }
                    )
                ),
            ),
        )
        .reset_index()
    )
    grouped["Uke"] = grouped["week_number"].apply(lambda value: f"Uke {int(value)}" if pd.notna(value) else "")
    grouped["Dato"] = grouped["published_date"].apply(format_date)
    grouped["published_date_sort"] = pd.to_datetime(grouped["published_date"], errors="coerce")
    grouped["Løp denne uken"] = grouped["lop"]
    grouped["Resultater"] = grouped["resultater"]
    grouped["Løpere"] = grouped["lopere"]

    summary = grouped[["published_date_sort", "Uke", "Dato", "Resultater", "Løpere", "Løp denne uken"]].copy()
    summary = summary.sort_values(["published_date_sort", "Uke"], ascending=[False, False], na_position="last")
    summary = summary.drop(columns=["published_date_sort"])
    return summary


def write_cards(ws, display_df: pd.DataFrame, raw_df: pd.DataFrame) -> None:
    cards = [
        ("Resultater", len(display_df)),
        ("Løpere", display_df["Navn"].nunique()),
        ("Løp", raw_df["event_name"].nunique()),
        ("Siste uke", f"Uke {int(raw_df['week_number'].max())}" if raw_df["week_number"].notna().any() else ""),
    ]

    start_col = 1
    for label, value in cards:
        ws.cell(row=4, column=start_col, value=label)
        ws.cell(row=5, column=start_col, value=value)
        for row in (4, 5):
            cell = ws.cell(row=row, column=start_col)
            cell.fill = CARD_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=4, column=start_col).font = Font(bold=True, color="7A1C1C")
        ws.cell(row=5, column=start_col).font = Font(bold=True, size=13)
        ws.column_dimensions[get_column_letter(start_col)].width = 18
        start_col += 2


def style_sheet_title(ws, title: str, subtitle: str, end_col: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    ws["A1"] = title
    ws["A2"] = subtitle
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 18


def write_dataframe_with_table(ws, df: pd.DataFrame, start_row: int, table_name: str) -> None:
    for col_idx, column_name in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=column_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_offset, row in enumerate(df.itertuples(index=False), 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=start_row + row_offset, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if col_idx in {1, 2, 5, 6, 7}:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")

    end_row = start_row + len(df)
    end_col_letter = get_column_letter(len(df.columns))
    table = Table(displayName=table_name, ref=f"A{start_row}:{end_col_letter}{end_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    widths = {
        1: 12,
        2: 14,
        3: 30,
        4: 24,
        5: 10,
        6: 16,
        7: 14,
        8: 14,
        9: 10,
        10: 34,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def build_workbook(display_df: pd.DataFrame, summary_df: pd.DataFrame, raw_df: pd.DataFrame) -> Workbook:
    wb = Workbook()

    ws_overview = wb.active
    ws_overview.title = "Delt oversikt"
    style_sheet_title(
        ws_overview,
        title="SK Vidar Langdistanse 2026",
        subtitle="Ukentlige resultater for klubbens langdistanseløpere",
        end_col=8,
    )
    write_cards(ws_overview, display_df, raw_df)
    write_dataframe_with_table(ws_overview, display_df, start_row=8, table_name="VidarLangdistanse")
    ws_overview.freeze_panes = "A9"
    ws_overview.sheet_view.showGridLines = False

    ws_summary = wb.create_sheet("Ukesammendrag")
    style_sheet_title(
        ws_summary,
        title="Ukesammendrag",
        subtitle="Kjapp oversikt over hvor mye som kom inn hver uke",
        end_col=5,
    )
    write_dataframe_with_table(ws_summary, summary_df, start_row=5, table_name="VidarUkesammendrag")
    ws_summary.freeze_panes = "A6"
    ws_summary.sheet_view.showGridLines = False
    ws_summary.column_dimensions["E"].width = 50

    ws_info = wb.create_sheet("Bruk")
    style_sheet_title(
        ws_info,
        title="Oppdatering",
        subtitle="Slik holder du den delte oversikten oppdatert",
        end_col=2,
    )
    instructions = [
        "1. Legg nye resultater inn i data/arbeidsfiler/weekly_results_2026.xlsx.",
        "2. Kjor scripts/build_shared_weekly_results_2026.py eller batch-filen i prosjektroten.",
        "3. Den delte filen lagres i data/delt_oversikt/SK Vidar Langdistanse 2026.xlsx.",
        "4. Inputfiler og stottefiler er samlet under data/ for enklere orden.",
    ]
    for row_idx, text in enumerate(instructions, 4):
        cell = ws_info.cell(row=row_idx, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = THIN_BORDER
    ws_info.column_dimensions["A"].width = 90
    ws_info.sheet_view.showGridLines = False

    return wb


def main() -> None:
    if not INPUT_FILE.exists():
        raise FileNotFoundError(f"Missing source workbook: {INPUT_FILE}")

    results_df = pd.read_excel(INPUT_FILE, sheet_name="results", engine="openpyxl")
    display_df = build_display_rows(results_df)
    summary_df = build_week_summary(results_df)
    workbook = build_workbook(display_df, summary_df, results_df)
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_FILE)

    print(f"Created shared workbook: {OUTPUT_FILE}")
    print(f"Rows in shared overview: {len(display_df)}")
    print(f"Week summary rows: {len(summary_df)}")


if __name__ == "__main__":
    main()

