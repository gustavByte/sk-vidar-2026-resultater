from __future__ import annotations

import argparse
import shutil
import tempfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from project_paths import WEEKLY_RESULTS_FILE
from result_taxonomy import classify_note_for_import, split_public_internal_note


RESULTS_SHEET = "results"
PUBLIC_NOTE_COLUMN = "public_note"
INTERNAL_NOTE_COLUMN = "internal_note"


@dataclass(frozen=True)
class MigrationSummary:
    rows: int
    changed_rows: int
    public_notes: int
    internal_notes: int
    backup_path: Path | None


def _backup_path(path: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%dT%H%M%S%f")
    return path.with_name(f"{path.stem}.pre-note-migration-{timestamp}{path.suffix}")


def migrate_workbook(path: Path = WEEKLY_RESULTS_FILE, *, create_backup: bool = True) -> MigrationSummary:
    """Add explicit note channels without changing or deleting the raw notes."""

    workbook = load_workbook(path)
    if RESULTS_SHEET not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"Missing sheet: {RESULTS_SHEET}")

    sheet = workbook[RESULTS_SHEET]
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    if "notes" not in headers:
        workbook.close()
        raise ValueError("Missing required notes column")

    schema_changed = False
    for column_name in (PUBLIC_NOTE_COLUMN, INTERNAL_NOTE_COLUMN):
        if column_name not in headers:
            headers.append(column_name)
            sheet.cell(row=1, column=len(headers), value=column_name)
            schema_changed = True

    header_map = {name: index + 1 for index, name in enumerate(headers)}
    changed_rows = 0
    public_notes = 0
    internal_notes = 0
    row_count = max(sheet.max_row - 1, 0)

    for row_index in range(2, sheet.max_row + 1):
        raw_note = sheet.cell(row=row_index, column=header_map["notes"]).value
        public_cell = sheet.cell(row=row_index, column=header_map[PUBLIC_NOTE_COLUMN])
        internal_cell = sheet.cell(row=row_index, column=header_map[INTERNAL_NOTE_COLUMN])

        existing_public = public_cell.value
        existing_internal = internal_cell.value
        if str(existing_public or "").strip() or str(existing_internal or "").strip():
            public_note, internal_note = split_public_internal_note(
                raw_note,
                existing_public,
                existing_internal,
            )
        else:
            public_note, internal_note = classify_note_for_import(raw_note)

        current_public = public_cell.value if public_cell.value is not None else ""
        current_internal = internal_cell.value if internal_cell.value is not None else ""
        if current_public != public_note or current_internal != internal_note:
            public_cell.value = public_note
            internal_cell.value = internal_note
            changed_rows += 1
        public_notes += bool(public_note)
        internal_notes += bool(internal_note)

    backup_path: Path | None = None
    if schema_changed or changed_rows:
        with tempfile.NamedTemporaryFile(suffix=path.suffix, dir=path.parent, delete=False) as handle:
            staged_path = Path(handle.name)
        try:
            workbook.save(staged_path)
            if create_backup:
                backup_path = _backup_path(path)
                shutil.copy2(path, backup_path)
            staged_path.replace(path)
        finally:
            staged_path.unlink(missing_ok=True)
    workbook.close()

    return MigrationSummary(
        rows=row_count,
        changed_rows=changed_rows,
        public_notes=public_notes,
        internal_notes=internal_notes,
        backup_path=backup_path,
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="Del rånotater i eksplisitte offentlige og interne felt.")
    parser.add_argument("--workbook", type=Path, default=WEEKLY_RESULTS_FILE)
    parser.add_argument("--no-backup", action="store_true")
    args = parser.parse_args()

    summary = migrate_workbook(args.workbook, create_backup=not args.no_backup)
    print(f"Rows checked: {summary.rows}")
    print(f"Rows changed: {summary.changed_rows}")
    print(f"Public notes: {summary.public_notes}")
    print(f"Internal notes: {summary.internal_notes}")
    if summary.backup_path:
        print(f"Backup: {summary.backup_path}")


if __name__ == "__main__":
    main()
