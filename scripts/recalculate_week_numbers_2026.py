from __future__ import annotations

from openpyxl import load_workbook
import pandas as pd
from project_paths import WEEKLY_RESULTS_FILE


WORKBOOK = WEEKLY_RESULTS_FILE
TARGET_SHEETS = ("results", "events", "review")


def iso_week(value: object) -> int | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    try:
        return int(pd.to_datetime(value).isocalendar().week)
    except Exception:
        return None


def main() -> None:
    if not WORKBOOK.exists():
        raise FileNotFoundError(f"Missing workbook: {WORKBOOK}")

    wb = load_workbook(WORKBOOK)
    total_updates = 0

    for sheet_name in TARGET_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        header_map = {str(value).strip(): idx + 1 for idx, value in enumerate(headers) if value is not None}
        if "published_date" not in header_map or "week_number" not in header_map:
            continue

        date_col = header_map["published_date"]
        week_col = header_map["week_number"]
        sheet_updates = 0

        for row_idx in range(2, ws.max_row + 1):
            date_value = ws.cell(row_idx, date_col).value
            recalculated = iso_week(date_value)
            if recalculated is None:
                continue

            current = ws.cell(row_idx, week_col).value
            current_int = None
            try:
                if current not in (None, ""):
                    current_int = int(current)
            except Exception:
                current_int = None

            if current_int != recalculated:
                ws.cell(row=row_idx, column=week_col, value=recalculated)
                sheet_updates += 1

        total_updates += sheet_updates
        print(f"{sheet_name}: updated {sheet_updates} rows")

    wb.save(WORKBOOK)
    print(f"Saved workbook: {WORKBOOK}")
    print(f"Total week number updates: {total_updates}")


if __name__ == "__main__":
    main()
