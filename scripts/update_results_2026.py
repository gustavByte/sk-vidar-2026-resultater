from __future__ import annotations

import argparse
import shutil
import subprocess
import sys
import tempfile
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from project_paths import (
    IMPORT_INBOX_DIR,
    IMPORT_MANIFEST_FILE,
    IMPORT_REVIEW_FILE,
    ROOT_DIR,
    WEEKLY_RESULTS_FILE,
)
from result_import import ImportCandidate, adapt_source, file_sha256, result_key
from spreadsheet_security import csv_safe_dataframe, set_openpyxl_literal_cell


RESULTS_SHEET = "results"
SUPPORTED_SUFFIXES = {".csv", ".txt", ".tsv", ".xlsx", ".xlsm"}
IDENTITY_SOURCE_COLUMNS = ("slack_user_id", "world_athletics_id", "source_system", "source_person_id")
NOTE_COLUMNS = ("public_note", "internal_note")


def imported_hashes() -> set[str]:
    if not IMPORT_MANIFEST_FILE.exists():
        return set()
    frame = pd.read_csv(IMPORT_MANIFEST_FILE, dtype=str).fillna("")
    return set(frame.loc[frame["status"].eq("imported"), "sha256"])


def scan_inbox() -> tuple[list[tuple[Path, str, list[ImportCandidate]]], list[dict[str, object]]]:
    IMPORT_INBOX_DIR.mkdir(parents=True, exist_ok=True)
    seen = imported_hashes()
    batches: list[tuple[Path, str, list[ImportCandidate]]] = []
    review_rows: list[dict[str, object]] = []

    for path in sorted(IMPORT_INBOX_DIR.iterdir()):
        if not path.is_file() or path.suffix.casefold() not in SUPPORTED_SUFFIXES:
            continue
        digest = file_sha256(path)
        if digest in seen:
            continue
        candidates = adapt_source(path)
        file_has_review = any(candidate.status != "ready" for candidate in candidates)
        if file_has_review:
            candidates = [
                ImportCandidate(
                    candidate.source_file,
                    candidate.source_row,
                    candidate.candidate_id,
                    "review" if candidate.status != "ready" else "held",
                    candidate.issues or ("Holdes tilbake fordi filen har andre kontrollfeil",),
                    candidate.row,
                )
                for candidate in candidates
            ]
        batches.append((path, digest, candidates))
        for candidate in candidates:
            if candidate.status == "ready":
                continue
            review_rows.append(
                {
                    "source_file": candidate.source_file,
                    "source_row": candidate.source_row,
                    "candidate_id": candidate.candidate_id,
                    "status": candidate.status,
                    "issues": "; ".join(candidate.issues),
                    **{key: value for key, value in candidate.row.items() if key != "internal_note"},
                }
            )
    return batches, review_rows


def write_review(rows: list[dict[str, object]]) -> None:
    IMPORT_REVIEW_FILE.parent.mkdir(parents=True, exist_ok=True)
    csv_safe_dataframe(pd.DataFrame(rows)).to_csv(IMPORT_REVIEW_FILE, index=False, encoding="utf-8-sig")


def existing_result_keys() -> set[tuple[str, ...]]:
    frame = pd.read_excel(WEEKLY_RESULTS_FILE, sheet_name=RESULTS_SHEET, engine="openpyxl").fillna("")
    return {result_key(row.to_dict()) for _, row in frame.iterrows()}


def workbook_row(candidate: ImportCandidate, columns: list[str]) -> list[object]:
    source = candidate.row
    values = {
        "published_date": source.get("published_date", ""),
        "week_number": source.get("week_number", ""),
        "event_name": source.get("event_name", ""),
        "distance": source.get("distance", ""),
        "category": source.get("class_name", ""),
        "athlete_name": source.get("athlete_name", ""),
        "slack_user_id": source.get("slack_user_id", ""),
        "world_athletics_id": source.get("world_athletics_id", ""),
        "source_system": source.get("source_system", ""),
        "source_person_id": source.get("source_person_id", ""),
        "result_time_raw": source.get("result_time_raw", ""),
        "result_time_normalized": source.get("result_time_raw", ""),
        "position": source.get("position", ""),
        "notes": source.get("notes", ""),
        "public_note": source.get("public_note", ""),
        "internal_note": source.get("internal_note", ""),
        "raw_entry": f"Importert fra {candidate.source_file}",
        "source_ts": f"adapter:{candidate.candidate_id}",
        "source_order": candidate.source_row,
        "WA Kjønn": source.get("gender", ""),
        "gender": source.get("gender", ""),
        "class_name": source.get("class_name", ""),
        "class_place": source.get("class_place", ""),
    }
    return [values.get(column, "") for column in columns]


def append_candidates(candidates: list[ImportCandidate]) -> tuple[int, int]:
    existing = existing_result_keys()
    unique: list[ImportCandidate] = []
    duplicates = 0
    for candidate in candidates:
        key = result_key(candidate.row)
        if key in existing:
            duplicates += 1
            continue
        existing.add(key)
        unique.append(candidate)

    if not unique:
        return 0, duplicates

    with tempfile.NamedTemporaryFile(suffix=".xlsx", dir=WEEKLY_RESULTS_FILE.parent, delete=False) as handle:
        temp_path = Path(handle.name)
    try:
        shutil.copy2(WEEKLY_RESULTS_FILE, temp_path)
        workbook = load_workbook(temp_path)
        sheet = workbook[RESULTS_SHEET]
        columns = [str(cell.value or "") for cell in sheet[1]]
        extension_columns = [
            column
            for column in (*NOTE_COLUMNS, *IDENTITY_SOURCE_COLUMNS)
            if column in NOTE_COLUMNS or any(candidate.row.get(column) for candidate in unique)
        ]
        for extension_column in extension_columns:
            if extension_column not in columns:
                columns.append(extension_column)
                set_openpyxl_literal_cell(sheet, 1, len(columns), extension_column)
        for candidate in unique:
            row_index = sheet.max_row + 1
            for column_index, value in enumerate(workbook_row(candidate, columns), 1):
                set_openpyxl_literal_cell(sheet, row_index, column_index, value)
        workbook.save(temp_path)
        temp_path.replace(WEEKLY_RESULTS_FILE)
    finally:
        temp_path.unlink(missing_ok=True)
    return len(unique), duplicates


def append_manifest(entries: list[dict[str, object]]) -> None:
    if not entries:
        return
    IMPORT_MANIFEST_FILE.parent.mkdir(parents=True, exist_ok=True)
    previous = pd.read_csv(IMPORT_MANIFEST_FILE, dtype=str).fillna("") if IMPORT_MANIFEST_FILE.exists() else pd.DataFrame()
    combined = pd.concat([previous, pd.DataFrame(entries)], ignore_index=True)
    csv_safe_dataframe(combined).to_csv(IMPORT_MANIFEST_FILE, index=False, encoding="utf-8-sig")


def run_build_pipeline() -> None:
    for script in (
        "scripts/recalculate_wa_points_2026.py",
        "scripts/build_shared_weekly_results_2026.py",
        "scripts/build_site_2026.py",
    ):
        subprocess.run([sys.executable, script], cwd=ROOT_DIR, check=True)


def main() -> None:
    parser = argparse.ArgumentParser(description="Trygg import og publiseringsbygg for SK Vidar-resultater.")
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--check", action="store_true", help="Kontroller inbox uten å endre arbeidsboken.")
    mode.add_argument("--build-only", action="store_true", help="Bygg eksisterende data uten å lese inbox.")
    args = parser.parse_args()

    if args.build_only:
        run_build_pipeline()
        return

    batches, review_rows = scan_inbox()
    write_review(review_rows)
    ready = [candidate for _, _, candidates in batches for candidate in candidates if candidate.status == "ready"]
    print(f"Inbox: {len(batches)} nye filer, {len(ready)} klare rader, {len(review_rows)} til kontroll")
    print(f"Kontrollrapport: {IMPORT_REVIEW_FILE}")

    if args.check:
        if review_rows:
            raise SystemExit(2)
        return

    backup = WEEKLY_RESULTS_FILE.with_suffix(".pre-import.xlsx")
    shutil.copy2(WEEKLY_RESULTS_FILE, backup)
    try:
        imported, duplicates = append_candidates(ready)
        now = datetime.now(timezone.utc).isoformat()
        manifest_entries = []
        for path, digest, candidates in batches:
            if any(candidate.status != "ready" for candidate in candidates):
                continue
            manifest_entries.append(
                {
                    "source_file": path.name,
                    "sha256": digest,
                    "rows": len(candidates),
                    "status": "imported",
                    "imported_at": now,
                }
            )
        run_build_pipeline()
        append_manifest(manifest_entries)
    except Exception:
        shutil.copy2(backup, WEEKLY_RESULTS_FILE)
        raise
    finally:
        backup.unlink(missing_ok=True)

    print(f"Importert: {imported}; duplikater hoppet over: {duplicates}")


if __name__ == "__main__":
    main()
