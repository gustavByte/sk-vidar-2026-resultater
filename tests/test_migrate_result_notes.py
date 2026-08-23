from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

from migrate_result_notes_2026 import migrate_workbook  # noqa: E402


def test_note_migration_preserves_raw_data_and_creates_backup(tmp_path: Path) -> None:
    path = tmp_path / "weekly_results_2026.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "results"
    sheet.append(["athlete_name", "notes"])
    sheet.append(["Kari", "PB. Sjekk Slack før publisering."])
    sheet.append(["Ola", ""])
    workbook.save(path)

    summary = migrate_workbook(path)

    assert summary.changed_rows == 1
    assert summary.backup_path is not None and summary.backup_path.exists()
    migrated = load_workbook(path, data_only=False)
    migrated_sheet = migrated["results"]
    headers = [cell.value for cell in migrated_sheet[1]]
    values = dict(zip(headers, [cell.value for cell in migrated_sheet[2]]))
    assert values["notes"] == "PB. Sjekk Slack før publisering."
    assert values["public_note"] == "PB"
    assert "Slack" in values["internal_note"]
    migrated.close()

    second = migrate_workbook(path)
    assert second.changed_rows == 0
    assert second.backup_path is None


def test_note_migration_persists_schema_for_empty_workbook(tmp_path: Path) -> None:
    path = tmp_path / "empty.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "results"
    sheet.append(["athlete_name", "notes"])
    workbook.save(path)

    summary = migrate_workbook(path, create_backup=False)

    assert summary.rows == 0
    migrated = load_workbook(path, data_only=False)
    headers = [cell.value for cell in migrated["results"][1]]
    assert headers == ["athlete_name", "notes", "public_note", "internal_note"]
    migrated.close()
