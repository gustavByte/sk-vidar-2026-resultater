from __future__ import annotations

import sys
import csv
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

from build_shared_weekly_results_2026 import build_display_rows, write_dataframe_with_table  # noqa: E402
from result_import import ImportCandidate  # noqa: E402
from spreadsheet_security import csv_safe_dataframe, is_formula_like_text  # noqa: E402
import update_results_2026  # noqa: E402


def test_formula_detection_covers_whitespace_and_control_prefixes() -> None:
    assert is_formula_like_text("=HYPERLINK(\"https://example.test\")")
    assert is_formula_like_text(" \t+1+1")
    assert is_formula_like_text("\x00@SUM(A1:A2)")
    assert is_formula_like_text("-2+3")
    assert not is_formula_like_text("Resultat = 12")
    assert not is_formula_like_text("'=-allerede-sikret")


def test_csv_export_neutralizes_only_formula_like_text() -> None:
    frame = pd.DataFrame(
        {
            "value": ["=1+1", "  @SUM(A1:A2)", "vanlig tekst", 42],
        }
    )

    secured = csv_safe_dataframe(frame)

    assert secured["value"].tolist() == ["'=1+1", "'  @SUM(A1:A2)", "vanlig tekst", 42]


def test_shared_workbook_stores_untrusted_formula_as_literal_text(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    write_dataframe_with_table(
        sheet,
        pd.DataFrame([{"Navn": "=WEBSERVICE(\"https://example.test\")"}]),
        start_row=1,
        table_name="SafeExport",
    )
    output = tmp_path / "safe.xlsx"
    workbook.save(output)

    reopened = load_workbook(output, data_only=False)
    cell = reopened.active["A2"]
    assert cell.value == "=WEBSERVICE(\"https://example.test\")"
    assert cell.data_type == "s"
    reopened.close()


def test_shared_export_rejects_unclassified_raw_note() -> None:
    frame = pd.DataFrame(
        [
            {
                "athlete_name": "Kari",
                "notes": "PB",
                "public_note": "",
                "internal_note": "",
            }
        ]
    )

    try:
        build_display_rows(frame)
    except ValueError as error:
        assert "uklassifiserte" in str(error)
    else:
        raise AssertionError("Unclassified note should stop the shared export")


def test_import_review_csv_neutralizes_formula_text(tmp_path: Path, monkeypatch) -> None:
    review_file = tmp_path / "review.csv"
    monkeypatch.setattr(update_results_2026, "IMPORT_REVIEW_FILE", review_file)

    update_results_2026.write_review([{"athlete_name": "\t=HYPERLINK(\"https://example.test\")"}])

    with review_file.open(encoding="utf-8-sig", newline="") as handle:
        row = next(csv.DictReader(handle))
    assert row["athlete_name"] == "'\t=HYPERLINK(\"https://example.test\")"


def test_working_workbook_keeps_formula_text_literal(tmp_path: Path, monkeypatch) -> None:
    workbook_path = tmp_path / "weekly_results_2026.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "results"
    sheet.append(["published_date", "event_name", "distance", "athlete_name", "result_time_raw"])
    workbook.save(workbook_path)
    monkeypatch.setattr(update_results_2026, "WEEKLY_RESULTS_FILE", workbook_path)

    candidate = ImportCandidate(
        source_file="source.csv",
        source_row=2,
        candidate_id="candidate-1",
        status="ready",
        issues=(),
        row={
            "published_date": "2026-08-23",
            "event_name": "Testløpet",
            "distance": "10 km",
            "athlete_name": "=WEBSERVICE(\"https://example.test\")",
            "result_time_raw": "35:00",
            "notes": "",
            "public_note": "",
            "internal_note": "",
        },
    )

    imported, duplicates = update_results_2026.append_candidates([candidate])

    assert (imported, duplicates) == (1, 0)
    reopened = load_workbook(workbook_path, data_only=False)
    result_sheet = reopened["results"]
    headers = [cell.value for cell in result_sheet[1]]
    athlete_cell = result_sheet.cell(row=2, column=headers.index("athlete_name") + 1)
    assert athlete_cell.value == "=WEBSERVICE(\"https://example.test\")"
    assert athlete_cell.data_type == "s"
    reopened.close()
